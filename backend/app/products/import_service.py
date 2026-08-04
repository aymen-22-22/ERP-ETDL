import contextlib
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.inventory.models import ProductStockSnapshot
from app.products.models import (
    Brand,
    Category,
    CategoryVariantScheme,
    Product,
    ProductStatus,
    ProductType,
    Unit,
)
from app.products.schemas import (
    ImportRowError,
    ImportSummary,
    ProductCreate,
    ProductRead,
    ProductUpdate,
)
from app.products.service import create_product, update_product
from app.products.variant_service import _sku_segment, build_name, build_sku
from app.shared.core.ids import generate_uuid7
from app.warehouses.models import Warehouse

# Column order shared by the template, the export, and the import parser --
# keeping the three in lockstep is what makes "export, tweak in Excel,
# re-import" a safe round trip instead of a column-shuffling trap.
TEMPLATE_COLUMNS = [
    ("name", "Name"),
    ("sku", "SKU"),
    ("barcode", "Barcode"),
    ("description", "Description"),
    ("price", "Price"),
    ("cost_price", "Cost price"),
    ("status", "Status"),
    ("category", "Category path (e.g. Lustre > Sous Lustre)"),
    ("colour", "Colour (leave blank for a simple product)"),
    ("brand", "Brand"),
    ("unit", "Unit"),
    ("warehouse", "Default warehouse"),
    ("stock_qty", "Stock qty (default warehouse)"),
]

_STATUS_VALUES = [s.value for s in ProductStatus]

_EXAMPLE_ROW = (
    "DTF Film Roll 30cm",
    "FLM-30-100",
    "6975012345678",
    "Premium DTF transfer film, 30cm x 100m roll",
    49.99,
    35.00,
    "active",
    "Films > Transfer Film",
    "",
    "DTF Pro",
    "Roll",
    "Main Warehouse",
    50,
)


async def _reference_rows(
    session: AsyncSession, tenant_id: UUID
) -> tuple[list[str], list[str], list[str], list[str]]:
    """Existing category paths / brand names / unit names / warehouse names
    for this tenant, so the template's Reference sheet shows exactly the
    values the import parser will recognise -- copy-paste instead of guess.
    """
    categories = list(
        (
            await session.scalars(
                select(Category).where(
                    Category.tenant_id == tenant_id, Category.deleted_at.is_(None)
                )
            )
        ).all()
    )
    by_id = {c.id: c for c in categories}

    def path_of(cat: Category) -> str:
        parts = [cat.name]
        cur = cat
        while cur.parent_id is not None and cur.parent_id in by_id:
            cur = by_id[cur.parent_id]
            parts.append(cur.name)
        return " > ".join(reversed(parts))

    category_paths = sorted(path_of(c) for c in categories)

    brands = list(
        (
            await session.scalars(
                select(Brand.name).where(Brand.tenant_id == tenant_id, Brand.deleted_at.is_(None))
            )
        ).all()
    )
    units = list(
        (
            await session.scalars(
                select(Unit.name).where(Unit.tenant_id == tenant_id, Unit.deleted_at.is_(None))
            )
        ).all()
    )
    warehouses = list(
        (
            await session.scalars(
                select(Warehouse.name).where(
                    Warehouse.tenant_id == tenant_id, Warehouse.deleted_at.is_(None)
                )
            )
        ).all()
    )
    return category_paths, sorted(brands), sorted(units), sorted(warehouses)


def _build_workbook(
    *,
    rows: list[tuple[object, ...]],
    category_paths: list[str],
    brands: list[str],
    units: list[str],
    warehouses: list[str],
    include_example: bool,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Products"
    ws.append([label for _, label in TEMPLATE_COLUMNS])
    if include_example and not rows:
        ws.append(list(_EXAMPLE_ROW))
    for row in rows:
        ws.append(list(row))

    status_col_letter = chr(ord("A") + [key for key, _ in TEMPLATE_COLUMNS].index("status"))
    status_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(_STATUS_VALUES)}"',
        allow_blank=True,
    )
    status_dv.error = "Must be one of: " + ", ".join(_STATUS_VALUES)
    ws.add_data_validation(status_dv)
    status_dv.add(f"{status_col_letter}2:{status_col_letter}1000")

    widths = [28, 18, 16, 32, 10, 12, 10, 30, 16, 16, 12, 20, 22]
    for i, width in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = width

    ref = wb.create_sheet("Reference")
    ref.append(["Existing categories", "Existing brands", "Existing units", "Existing warehouses"])
    for i in range(max(len(category_paths), len(brands), len(units), len(warehouses))):
        ref.append(
            [
                category_paths[i] if i < len(category_paths) else None,
                brands[i] if i < len(brands) else None,
                units[i] if i < len(units) else None,
                warehouses[i] if i < len(warehouses) else None,
            ]
        )
    for col in "ABCD":
        ref.column_dimensions[col].width = 26

    instructions = wb.create_sheet("Instructions")
    for line in (
        "Fill in the Products sheet and upload it via Import.",
        "",
        "Name, SKU, and Price are required. Everything else is optional.",
        "SKU is the matching key: an existing SKU updates that product; a new"
        " SKU creates one. Nothing is ever deleted by importing.",
        "",
        "Status: draft, active, or archived (defaults to active if left blank).",
        "Category path: use > to separate levels, e.g. Lustre > Sous Lustre."
        " Unknown categories/brands are created automatically.",
        "Colour: fill this in to add a colour of an existing product family."
        " Give the SAME Name and Category as another row -- that row becomes"
        " the family, and its formula (when the category has one) computes"
        " the final Name/SKU for you, so what you type in Name/SKU for a"
        " colour row is a starting point, not final. Leave blank for an"
        " ordinary simple product.",
        "Unit and Default warehouse must already exist in the app (see the"
        " Reference sheet) -- these are not auto-created, since a typo would"
        " otherwise silently create a duplicate warehouse.",
        "Stock qty only applies to new products (it seeds their opening stock"
        " at the default warehouse). It's ignored when updating an existing"
        " product -- adjust stock from the product page or a transfer instead.",
        "",
        "See the Reference sheet for the exact spelling of your existing"
        " categories, brands, units, and warehouses.",
    ):
        instructions.append([line])
    instructions.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def generate_template(session: AsyncSession, tenant_id: UUID) -> bytes:
    category_paths, brands, units, warehouses = await _reference_rows(session, tenant_id)
    return _build_workbook(
        rows=[],
        category_paths=category_paths,
        brands=brands,
        units=units,
        warehouses=warehouses,
        include_example=True,
    )


async def export_products(session: AsyncSession, tenant_id: UUID) -> bytes:
    products = list(
        (
            await session.scalars(
                select(Product)
                .where(Product.tenant_id == tenant_id, Product.deleted_at.is_(None))
                .order_by(Product.name)
            )
        ).all()
    )
    category_paths, brands, units, warehouses = await _reference_rows(session, tenant_id)

    categories_by_id = {
        c.id: c
        for c in (
            await session.scalars(
                select(Category).where(
                    Category.tenant_id == tenant_id, Category.deleted_at.is_(None)
                )
            )
        ).all()
    }
    brands_by_id = {
        b.id: b.name
        for b in (
            await session.scalars(
                select(Brand).where(Brand.tenant_id == tenant_id, Brand.deleted_at.is_(None))
            )
        ).all()
    }
    units_by_id = {
        u.id: u.name
        for u in (
            await session.scalars(
                select(Unit).where(Unit.tenant_id == tenant_id, Unit.deleted_at.is_(None))
            )
        ).all()
    }
    warehouses_by_id = {
        w.id: w.name
        for w in (
            await session.scalars(
                select(Warehouse).where(
                    Warehouse.tenant_id == tenant_id, Warehouse.deleted_at.is_(None)
                )
            )
        ).all()
    }
    stock_by_key = {
        (s.product_id, s.warehouse_id): s.quantity_on_hand
        for s in (
            await session.scalars(
                select(ProductStockSnapshot).where(ProductStockSnapshot.tenant_id == tenant_id)
            )
        ).all()
    }
    color_key_by_category = {
        s.category_id: s.color_key
        for s in (
            await session.scalars(
                select(CategoryVariantScheme).where(
                    CategoryVariantScheme.tenant_id == tenant_id,
                    CategoryVariantScheme.deleted_at.is_(None),
                )
            )
        ).all()
        if s.color_key
    }

    def category_path(cat_id: UUID | None) -> str | None:
        if cat_id is None or cat_id not in categories_by_id:
            return None
        parts = []
        cur: Category | None = categories_by_id[cat_id]
        while cur is not None:
            parts.append(cur.name)
            cur = categories_by_id.get(cur.parent_id) if cur.parent_id else None
        return " > ".join(reversed(parts))

    rows: list[tuple[object, ...]] = []
    for p in products:
        stock_qty = (
            stock_by_key.get((p.id, p.default_warehouse_id)) if p.default_warehouse_id else None
        )
        colour = None
        if p.product_type == ProductType.VARIANT and p.category_id is not None:
            color_key = color_key_by_category.get(p.category_id)
            if color_key:
                colour = (p.attributes or {}).get(color_key) or None
        rows.append(
            (
                p.name,
                p.sku,
                p.barcode,
                p.description,
                float(p.price),
                float(p.cost_price) if p.cost_price is not None else None,
                p.status.value,
                category_path(p.category_id),
                colour,
                brands_by_id.get(p.brand_id) if p.brand_id else None,
                units_by_id.get(p.unit_id) if p.unit_id else None,
                warehouses_by_id.get(p.default_warehouse_id) if p.default_warehouse_id else None,
                stock_qty,
            )
        )

    return _build_workbook(
        rows=rows,
        category_paths=category_paths,
        brands=brands,
        units=units,
        warehouses=warehouses,
        include_example=False,
    )


async def _get_or_create_category(session: AsyncSession, tenant_id: UUID, path: str) -> UUID | None:
    parts = [p.strip() for p in path.split(">") if p.strip()]
    parent_id: UUID | None = None
    category_id: UUID | None = None
    for part in parts:
        cat = await session.scalar(
            select(Category).where(
                Category.tenant_id == tenant_id,
                Category.name == part,
                Category.parent_id == parent_id,
                Category.deleted_at.is_(None),
            )
        )
        if cat is None:
            cat = Category(id=generate_uuid7(), tenant_id=tenant_id, name=part, parent_id=parent_id)
            session.add(cat)
            await session.flush()
        parent_id = cat.id
        category_id = cat.id
    return category_id


async def _get_or_create_brand(session: AsyncSession, tenant_id: UUID, name: str) -> UUID:
    brand = await session.scalar(
        select(Brand).where(
            Brand.tenant_id == tenant_id, Brand.name == name, Brand.deleted_at.is_(None)
        )
    )
    if brand is None:
        brand = Brand(id=generate_uuid7(), tenant_id=tenant_id, name=name)
        session.add(brand)
        await session.flush()
    return brand.id


async def _find_unit(session: AsyncSession, tenant_id: UUID, text: str) -> UUID | None:
    return await session.scalar(
        select(Unit.id).where(
            Unit.tenant_id == tenant_id,
            Unit.deleted_at.is_(None),
            (Unit.name.ilike(text)) | (Unit.abbreviation.ilike(text)),
        )
    )


async def _find_warehouse(session: AsyncSession, tenant_id: UUID, name: str) -> UUID | None:
    return await session.scalar(
        select(Warehouse.id).where(
            Warehouse.tenant_id == tenant_id,
            Warehouse.deleted_at.is_(None),
            Warehouse.name.ilike(name),
        )
    )


def _parse_decimal(raw: object) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        return None
    with contextlib.suppress(InvalidOperation, ValueError):
        return Decimal(str(raw).replace(",", "").strip())
    return None


def _parse_int(raw: object) -> int | None:
    if raw is None or str(raw).strip() == "":
        return None
    with contextlib.suppress(ValueError):
        return int(float(str(raw).replace(",", "").strip()))
    return None


async def import_products(
    session: AsyncSession,
    tenant_id: UUID,
    file_bytes: bytes,
) -> ImportSummary:
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active
    if ws is None:
        raise ValueError("Workbook has no sheets")

    created: list[ProductRead] = []
    updated: list[ProductRead] = []
    errors: list[ImportRowError] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        (
            name_raw,
            sku_raw,
            barcode_raw,
            description_raw,
            price_raw,
            cost_raw,
            status_raw,
            category_raw,
            colour_raw,
            brand_raw,
            unit_raw,
            warehouse_raw,
            stock_raw,
        ) = (list(row) + [None] * len(TEMPLATE_COLUMNS))[: len(TEMPLATE_COLUMNS)]

        name = str(name_raw or "").strip()
        sku = str(sku_raw or "").strip()
        if not name or not sku:
            errors.append(ImportRowError(row=row_idx, message="Name and SKU are required"))
            continue

        price = _parse_decimal(price_raw)
        if price is None or price <= 0:
            errors.append(ImportRowError(row=row_idx, message="Price must be a positive number"))
            continue

        try:
            cost_price = _parse_decimal(cost_raw)
            barcode = str(barcode_raw or "").strip() or None
            description = str(description_raw or "").strip() or None

            status = ProductStatus.ACTIVE
            status_text = str(status_raw or "").strip().lower()
            if status_text:
                try:
                    status = ProductStatus(status_text)
                except ValueError:
                    errors.append(
                        ImportRowError(
                            row=row_idx,
                            message=f"Unknown status '{status_text}', defaulted to active",
                        )
                    )

            category_id = None
            category_text = str(category_raw or "").strip()
            if category_text:
                category_id = await _get_or_create_category(session, tenant_id, category_text)

            colour = str(colour_raw or "").strip()
            product_type = ProductType.SIMPLE
            attributes: dict[str, str] = {}
            if colour and category_id is not None:
                product_type = ProductType.VARIANT
                scheme = await session.scalar(
                    select(CategoryVariantScheme).where(
                        CategoryVariantScheme.tenant_id == tenant_id,
                        CategoryVariantScheme.category_id == category_id,
                        CategoryVariantScheme.deleted_at.is_(None),
                    )
                )
                # The row sharing this Name + Category is the rest of the
                # family -- its attributes seed this colour's, same as
                # "Add colour" in the UI.
                sibling = await session.scalar(
                    select(Product)
                    .where(
                        Product.tenant_id == tenant_id,
                        Product.category_id == category_id,
                        Product.name == name,
                        Product.deleted_at.is_(None),
                    )
                    .order_by(Product.created_at)
                )
                if sibling is not None:
                    attributes = dict(sibling.attributes or {})
                if scheme is not None and scheme.color_key:
                    attributes[scheme.color_key] = colour
                elif scheme is not None:
                    attributes["color"] = colour

                if scheme is not None:
                    structural_keys = [
                        key for key in scheme.attribute_keys if key != scheme.color_key
                    ]
                    has_full_structure = all(
                        str(attributes.get(key, "")).strip() for key in structural_keys
                    )
                    if has_full_structure:
                        name = build_name(
                            scheme.base_name, scheme.attribute_keys, attributes, scheme.color_key
                        )
                        sku = build_sku(scheme.sku_prefix, scheme.attribute_keys, attributes)
                    elif sibling is not None:
                        name = sibling.name
                        color_segment = _sku_segment(colour)
                        sku = f"{sibling.sku}-{color_segment}" if color_segment else sibling.sku
                elif sibling is not None:
                    name = sibling.name
                    color_segment = _sku_segment(colour)
                    sku = f"{sibling.sku}-{color_segment}" if color_segment else sibling.sku

            brand_id = None
            brand_text = str(brand_raw or "").strip()
            if brand_text:
                brand_id = await _get_or_create_brand(session, tenant_id, brand_text)

            unit_id = None
            unit_text = str(unit_raw or "").strip()
            if unit_text:
                unit_id = await _find_unit(session, tenant_id, unit_text)
                if unit_id is None:
                    errors.append(
                        ImportRowError(
                            row=row_idx, message=f"Unknown unit '{unit_text}', left unset"
                        )
                    )

            warehouse_id = None
            warehouse_text = str(warehouse_raw or "").strip()
            if warehouse_text:
                warehouse_id = await _find_warehouse(session, tenant_id, warehouse_text)
                if warehouse_id is None:
                    errors.append(
                        ImportRowError(
                            row=row_idx,
                            message=f"Unknown warehouse '{warehouse_text}', left unset",
                        )
                    )

            existing = await session.scalar(
                select(Product).where(
                    Product.tenant_id == tenant_id,
                    Product.sku == sku,
                    Product.deleted_at.is_(None),
                )
            )

            if existing is not None:
                update_data = ProductUpdate(
                    name=name,
                    sku=sku,
                    barcode=barcode,
                    description=description,
                    price=price,
                    cost_price=cost_price,
                    status=status,
                    category_id=category_id,
                    brand_id=brand_id,
                    unit_id=unit_id,
                    default_warehouse_id=warehouse_id,
                )
                product = await update_product(session, tenant_id, existing.id, update_data)
                updated.append(ProductRead.model_validate(product))
            else:
                stock_qty = _parse_int(stock_raw)
                create_data = ProductCreate(
                    id=generate_uuid7(),
                    name=name,
                    sku=sku,
                    barcode=barcode,
                    description=description,
                    price=price,
                    cost_price=cost_price,
                    status=status,
                    product_type=product_type,
                    attributes=attributes,
                    category_id=category_id,
                    brand_id=brand_id,
                    unit_id=unit_id,
                    default_warehouse_id=warehouse_id,
                    initial_stock=stock_qty if warehouse_id else None,
                )
                product = await create_product(session, tenant_id, create_data)
                created.append(ProductRead.model_validate(product))
        except Exception as exc:  # noqa: BLE001 -- one bad row must not sink the batch
            await session.rollback()
            errors.append(ImportRowError(row=row_idx, message=str(exc) or "Import failed"))

    return ImportSummary(created=created, updated=updated, errors=errors)
